"""
Excel 读取器 —— 打开 .xlsx，解析 Sheet 结构，返回结构化数据。
"""
import os
from typing import Optional

import openpyxl

from engine.type_system import parse_type_definition, parse_value, FieldType, TypeKind


class FieldDef:
    """字段定义"""
    def __init__(self, note: str, name: str, type_def: FieldType, col_index: int):
        self.note = note          # 备注（中文说明）
        self.name = name          # 导出字段名
        self.type_def = type_def  # 类型定义
        self.col_index = col_index  # Excel 列号 (1-based)


class TableData:
    """单张表（一个 Sheet）的完整数据"""
    def __init__(self):
        self.export_name: str = ""          # 导出文件名（不含 .lua）
        self.source_label: str = ""         # 来源标识 "文件/Sheet"
        self.fields: list[FieldDef] = []    # 字段列表
        self.rows: list[dict] = []          # 数据行列表 [{field_name: value}, ...]
        self.has_id: bool = False
        self.has_subid: bool = False


def read_excel(file_path: str) -> list[TableData]:
    """读取单个 Excel 文件的所有 Sheet，返回 TableData 列表。"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    tables: list[TableData] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        table = _parse_sheet(ws, file_path, sheet_name)
        if table is not None:
            tables.append(table)

    wb.close()
    return tables


def _parse_sheet(ws, file_path: str, sheet_name: str) -> Optional[TableData]:
    """解析单个 Sheet。若 A1 无有效表名则返回 None。"""
    a1 = ws.cell(1, 1).value
    if a1 is None or str(a1).strip() == "":
        return None

    a1_str = str(a1).strip()

    # 解析 A1: "别名:导出文件名"
    if ':' not in a1_str:
        return None
    parts = a1_str.split(':', 1)
    export_name = parts[1].strip()
    if not export_name:
        return None

    table = TableData()
    table.export_name = export_name
    table.source_label = f"{os.path.basename(file_path)}/{sheet_name}"

    max_col = ws.max_column
    if max_col is None:
        max_col = 1

    # ---- 解析 Row1: 字段名（B列起） ----
    field_names_seen: set[str] = set()
    for col in range(2, max_col + 1):
        cell_val = ws.cell(1, col).value
        if cell_val is None:
            # 空列跳过（但继续检查后续列）
            continue
        raw = str(cell_val).strip()
        if not raw:
            continue

        # 解析 "备注|字段名"
        note, name = _parse_field_header(raw, file_path, sheet_name, col)

        # 重复字段名检查
        if name in field_names_seen:
            raise ValueError(
                f"[{os.path.basename(file_path)}/{sheet_name}] "
                f"第1行第{_col_letter(col)}列: 字段名 \"{name}\" 重复定义，不允许重复"
            )
        field_names_seen.add(name)

        # 解析 Row2: 类型定义
        type_raw = ws.cell(2, col).value
        if type_raw is None:
            raise ValueError(
                f"[{os.path.basename(file_path)}/{sheet_name}] "
                f"第2行第{_col_letter(col)}列: 缺少类型定义"
            )
        try:
            type_def = parse_type_definition(str(type_raw).strip())
        except ValueError as e:
            raise ValueError(
                f"[{os.path.basename(file_path)}/{sheet_name}] "
                f"第2行第{_col_letter(col)}列: {e}"
            )

        field_def = FieldDef(note=note, name=name, type_def=type_def, col_index=col)
        table.fields.append(field_def)

    if not table.fields:
        return table  # 无字段定义，返回空表

    # 检查 ID 字段
    id_field = _find_field(table.fields, "ID")
    if id_field is None:
        raise ValueError(
            f"[{os.path.basename(file_path)}/{sheet_name}] "
            f"缺少 ID 字段，所有表必须包含 ID 列"
        )
    table.has_id = True

    # 检查 SubID 字段
    subid_field = _find_field(table.fields, "SubID")
    table.has_subid = subid_field is not None

    # ---- 确定数据行范围 ----
    # 从第3行起，找到 A 列最后一个非 0 的行
    last_data_row = _find_last_data_row(ws, max_col)
    if last_data_row < 3:
        return table

    # ---- 解析数据行 ----
    for row in range(3, last_data_row + 1):
        # 检查 A 列导出标记
        export_flag = ws.cell(row, 1).value
        if export_flag is None:
            continue
        try:
            flag = int(export_flag)
        except (ValueError, TypeError):
            continue
        if flag != 1:
            continue

        # 解析本行各字段
        row_data: dict = {}
        for fd in table.fields:
            raw_val = ws.cell(row, fd.col_index).value
            if raw_val is None:
                row_data[fd.name] = None
                continue

            raw_str = str(raw_val).strip()
            if not raw_str:
                row_data[fd.name] = None
                continue

            try:
                parsed = parse_value(raw_str, fd.type_def)
                row_data[fd.name] = parsed
            except ValueError as e:
                raise ValueError(
                    f"[{os.path.basename(file_path)}/{sheet_name}] "
                    f"第{row}行第{_col_letter(fd.col_index)}列 "
                    f"({fd.note}|{fd.name}): {e}\n"
                    f"  原始值: {raw_str}"
                )

        # 检查 ID 不能为空
        if row_data.get("ID") is None:
            raise ValueError(
                f"[{os.path.basename(file_path)}/{sheet_name}] "
                f"第{row}行: ID 不能为空"
            )

        table.rows.append(row_data)

    return table


def _parse_field_header(raw: str, file_path: str, sheet_name: str, col: int) -> tuple[str, str]:
    """解析 "备注|字段名"，返回 (note, name)。"""
    if '|' not in raw:
        # 无备注，整个作为字段名
        return "", raw
    parts = raw.split('|', 1)
    note = parts[0].strip()
    name = parts[1].strip()
    if not name:
        raise ValueError(
            f"[{os.path.basename(file_path)}/{sheet_name}] "
            f"第1行第{_col_letter(col)}列: 字段名为空"
        )
    return note, name


def _find_field(fields: list[FieldDef], name: str) -> Optional[FieldDef]:
    """查找指定名称的字段（大小写敏感）。"""
    for fd in fields:
        if fd.name == name:
            return fd
    return None


def _find_last_data_row(ws, max_col: int) -> int:
    """从后往前扫描，找到 A 列最后一个不为 0 的行（从第3行起）。
    这样在数据末尾有空行时能快速定位，避免扫描大量空行。
    """
    for row in range(ws.max_row, 2, -1):
        val = ws.cell(row, 1).value
        if val is None:
            continue
        try:
            if int(val) != 0:
                return row
        except (ValueError, TypeError):
            continue
    return 2  # 无有效数据行


def _col_letter(col_index: int) -> str:
    """将 1-based 列号转为 Excel 列字母。如 1->A, 2->B, 27->AA"""
    result = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        result = chr(65 + remainder) + result
    return result
