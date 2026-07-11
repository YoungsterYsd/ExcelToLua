"""
错误场景测试 —— 验证各种异常情况的错误提示是否清晰准确。
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from engine.excel_reader import read_excel
from engine.validator import validate_table, ValidationError
from engine.merger import merge_tables


def create_error_excel(cases: list[dict]) -> str:
    """创建包含错误场景的 Excel 文件，返回文件路径。"""
    wb = openpyxl.Workbook()
    
    for i, case in enumerate(cases):
        if i == 0:
            ws = wb.active
            ws.title = case.get("sheet_name", f"Sheet{i+1}")
        else:
            ws = wb.create_sheet(case.get("sheet_name", f"Sheet{i+1}"))
        
        name = case.get("a1", "测试:testError")
        ws.cell(1, 1, name)
        
        headers = case.get("headers", [])
        for j, h in enumerate(headers):
            ws.cell(1, j + 2, h)
        
        types_row = case.get("types", [])
        ws.cell(2, 1, "是否导出")
        for j, t in enumerate(types_row):
            ws.cell(2, j + 2, t)
        
        data_rows = case.get("data", [])
        for r, row_data in enumerate(data_rows):
            for c, val in enumerate(row_data):
                ws.cell(r + 3, c + 1, val)
    
    # 保存到临时文件
    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="test_error_")
    os.close(fd)
    wb.save(path)
    return path


def run_error_tests():
    """运行所有错误场景测试。"""
    print("=" * 60)
    print("错误场景测试")
    print("=" * 60)

    # ---- 测试1: 重复字段名 ----
    print("\n[测试1] 重复字段名...")
    path1 = create_error_excel([{
        "sheet_name": "重复字段",
        "headers": ["识别ID|ID", "值A|Value", "值A|Value"],  # Value 重复
        "types": ["Int", "Int", "Int"],
        "data": [[1, 1001, 10, 20]],
    }])
    try:
        read_excel(path1)
        print("  ❌ 未检测到重复字段名！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path1)

    # ---- 测试2: 缺少 ID 字段 ----
    print("\n[测试2] 缺少 ID 字段...")
    path2 = create_error_excel([{
        "sheet_name": "缺ID",
        "headers": ["名称|Name", "数值|Value"],
        "types": ["String", "Int"],
        "data": [[1, "测试", 10]],
    }])
    try:
        tables = read_excel(path2)
        print("  ❌ 未检测到缺少 ID！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path2)

    # ---- 测试3: ID 重复但无 SubID ----
    print("\n[测试3] ID 重复但无 SubID 字段...")
    path3 = create_error_excel([{
        "sheet_name": "ID重复无SubID",
        "headers": ["识别ID|ID", "名称|Name"],
        "types": ["Int", "String"],
        "data": [[1, 1001, "数据A"], [1, 1001, "数据B"]],  # ID=1001 重复
    }])
    try:
        tables = read_excel(path3)
        merged = merge_tables(tables)
        for name, m in merged.items():
            validate_table(m)
        print("  ❌ 未检测到 ID 重复！")
    except (ValidationError, Exception) as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path3)

    # ---- 测试4: Const 值不在定义范围 ----
    print("\n[测试4] Const 值不在定义范围...")
    path4 = create_error_excel([{
        "sheet_name": "Const错误",
        "headers": ["识别ID|ID", "类型|ItemType"],
        "types": ["Int", "Const(Sword=1,Shield=2)"],
        "data": [[1, 1001, "Bow"]],  # Bow 未定义
    }])
    try:
        read_excel(path4)
        print("  ❌ 未检测到 Const 错误！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path4)

    # ---- 测试5: Int 类型填写了字符串 ----
    print("\n[测试5] Int 类型填写了非数字...")
    path5 = create_error_excel([{
        "sheet_name": "类型错误",
        "headers": ["识别ID|ID", "数值|Value"],
        "types": ["Int", "Int"],
        "data": [[1, 1001, "hello"]],  # hello 不能转为 Int
    }])
    try:
        read_excel(path5)
        print("  ❌ 未检测到类型错误！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path5)

    # ---- 测试6: Dic 格式错误（用 : 而非 =） ----
    print("\n[测试6] Dic 定义使用 : 分隔符...")
    path6 = create_error_excel([{
        "sheet_name": "Dic格式错",
        "headers": ["识别ID|ID", "字典|TestDic"],
        "types": ["Int", "Dic(String:Int)"],  # 应该用 = 而非 :
        "data": [[1, 1001, "{k=1}"]],
    }])
    try:
        read_excel(path6)
        print("  ❌ 未检测到 Dic 格式错误！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path6)

    # ---- 测试7: List 格式错误 ----
    print("\n[测试7] List 用 [] 而非 {} 包裹...")
    path7 = create_error_excel([{
        "sheet_name": "List格式错",
        "headers": ["识别ID|ID", "列表|TestList"],
        "types": ["Int", "List(Int)"],
        "data": [[1, 1001, "[1,2,3]"]],  # List 应该用 {}
    }])
    try:
        read_excel(path7)
        print("  ❌ 未检测到 List 格式错误！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path7)

    # ---- 测试8: 嵌套 List 类型不匹配 ----
    print("\n[测试8] List(List(Int)) 内层填写字符串...")
    path8 = create_error_excel([{
        "sheet_name": "嵌套List错误",
        "headers": ["识别ID|ID", "嵌套|Nested"],
        "types": ["Int", "List(List(Int))"],
        "data": [[1, 1001, "{{abc,def},{ghi,jkl}}"]],  # 字符串而非 Int
    }])
    try:
        read_excel(path8)
        print("  ❌ 未检测到嵌套 List 类型错误！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:200]}...")
    finally:
        os.unlink(path8)

    # ---- 测试9: Dic 数据键值对缺少 = ----
    print("\n[测试9] Dic 数据键值对用 : 分隔...")
    path9 = create_error_excel([{
        "sheet_name": "Dic数据错",
        "headers": ["识别ID|ID", "字典|TestDic"],
        "types": ["Int", "Dic(String=Int)"],
        "data": [[1, 1001, "{key:123}"]],  # 应该用 = 而非 :
    }])
    try:
        read_excel(path9)
        print("  ❌ 未检测到 Dic 数据格式错误！")
    except Exception as e:
        print(f"  ✅ 正确捕获: {str(e)[:120]}...")
    finally:
        os.unlink(path9)

    print(f"\n{'=' * 60}")
    print("错误场景测试完成")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    run_error_tests()
