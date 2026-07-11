"""
测试配置创建器 —— 自动生成包含各种场景的测试 Excel 文件。
覆盖：基础类型、Array、Dic、List、嵌套 List、Ignore、ID 唯一、ID+SubID、多表合并
"""
import os
import openpyxl


def create_test_excel(output_path: str):
    """创建用于测试的 Excel 文件。"""
    wb = openpyxl.Workbook()

    # ====== Sheet 1: 基础类型测试 (testConfig) ======
    ws1 = wb.active
    ws1.title = "基础类型"

    # Row 1: 表名 + 字段定义
    ws1.cell(1, 1, "基础配置表:testConfig")
    ws1.cell(1, 2, "识别ID|ID")
    ws1.cell(1, 3, "测试用Int|TestInt")
    ws1.cell(1, 4, "测试用Float|TestFloat")
    ws1.cell(1, 5, "测试用String|TestString")
    ws1.cell(1, 6, "测试用Const|TestConst")
    ws1.cell(1, 7, "测试用Ignore|TestIgnore")
    ws1.cell(1, 8, "测试用Array|TestArray")
    ws1.cell(1, 9, "测试用Dic|TestDic")
    ws1.cell(1, 10, "测试用ListInt|TestIntList")
    ws1.cell(1, 11, "测试用ListString|TestStringList")
    ws1.cell(1, 12, "嵌套List|NestedList")

    # Row 2: 类型定义
    ws1.cell(2, 1, "是否导出(1=导出,0/空=跳过)")
    ws1.cell(2, 2, "Int")
    ws1.cell(2, 3, "Int")
    ws1.cell(2, 4, "Float")
    ws1.cell(2, 5, "String")
    ws1.cell(2, 6, "Const(TypeA=1,TypeB=2,TypeC=3)")
    ws1.cell(2, 7, "Ignore")
    ws1.cell(2, 8, "Array")
    ws1.cell(2, 9, "Dic(String=Int)")
    ws1.cell(2, 10, "List(Int)")
    ws1.cell(2, 11, "List(String)")
    ws1.cell(2, 12, "List(List(Int))")

    # Row 3+: 数据
    data1 = [
        [1, 1001, 10, 3.14, "新手剑", "TypeA", "忽略文本", "[10,20,30]", "{攻击=15,防御=5}", "{1,2,3}", "{铁,铜,银}", "{{1,2},{3,4}}"],
        [1, 1002, 20, 2.71, "铁盾",   "TypeB", "忽略文本2","[5,15]",      "{生命=100}",        "{4,5}",   "{金,银}",     "{{5,6}}"],
        [0, 1003, 30, 1.41, "跳过行", "TypeC", "不该出现",  "[1]",         "{速度=10}",         "{7}",     "{铂}",        "{{7,8},{9,10}}"],
        [1, 1004, 40, 0.01, "皮甲",   "TypeC", "忽略文本3","[100,200,300]","{魔抗=30,护甲=20}","{10,20,30}","{龙,凤}",    "{{100,200}}"],
    ]
    for i, row_data in enumerate(data1):
        row_num = 3 + i
        for j, val in enumerate(row_data):
            ws1.cell(row_num, j + 1, val)

    # ====== Sheet 2: 分表合并测试 (testConfig) ======
    ws2 = wb.create_sheet("分表补充")

    ws2.cell(1, 1, "补充表:testConfig")
    ws2.cell(1, 2, "识别ID|ID")
    ws2.cell(1, 3, "测试用Int|TestInt")
    ws2.cell(1, 4, "新增字段|ExtraField")     # 仅在分表中存在的字段
    ws2.cell(1, 5, "测试用String|TestString")

    ws2.cell(2, 1, "是否导出")
    ws2.cell(2, 2, "Int")
    ws2.cell(2, 3, "Int")
    ws2.cell(2, 4, "String")
    ws2.cell(2, 5, "String")

    data2 = [
        [1, 2001, 50, "额外数据A", "分表字符串A"],
        [1, 2002, 60, "额外数据B", "分表字符串B"],
    ]
    for i, row_data in enumerate(data2):
        row_num = 3 + i
        for j, val in enumerate(row_data):
            ws2.cell(row_num, j + 1, val)

    # ====== Sheet 3: ID+SubID 嵌套测试 ======
    ws3 = wb.create_sheet("嵌套表")

    ws3.cell(1, 1, "嵌套配置:testNested")
    ws3.cell(1, 2, "识别ID|ID")
    ws3.cell(1, 3, "子ID|SubID")
    ws3.cell(1, 4, "名称|Name")
    ws3.cell(1, 5, "等级|Level")
    ws3.cell(1, 6, "攻击力|Atk")

    ws3.cell(2, 1, "是否导出")
    ws3.cell(2, 2, "Int")
    ws3.cell(2, 3, "Int")
    ws3.cell(2, 4, "String")
    ws3.cell(2, 5, "Int")
    ws3.cell(2, 6, "Int")

    data3 = [
        [1, 1001, 1, "新手剑",   1, 15],
        [1, 1001, 2, "新手剑+1", 2, 20],
        [1, 1001, 3, "新手剑+2", 3, 28],
        [1, 1002, 1, "铁盾",     1, 0],
        [1, 1002, 2, "铁盾+1",   2, 0],
        [1, 1003, 1, "皮甲",     1, 0],
    ]
    for i, row_data in enumerate(data3):
        row_num = 3 + i
        for j, val in enumerate(row_data):
            ws3.cell(row_num, j + 1, val)

    # ====== Sheet 4: String ID 测试 ======
    ws4 = wb.create_sheet("字符串ID")

    ws4.cell(1, 1, "字符串ID表:testStringId")
    ws4.cell(1, 2, "识别ID|ID")
    ws4.cell(1, 3, "名称|Name")
    ws4.cell(1, 4, "数值|Value")

    ws4.cell(2, 1, "是否导出")
    ws4.cell(2, 2, "String")
    ws4.cell(2, 3, "String")
    ws4.cell(2, 4, "Int")

    data4 = [
        [1, "ITEM_SWORD",  "铁剑", 100],
        [1, "ITEM_SHIELD", "铁盾", 50],
        [1, "ITEM_ARMOR",  "皮甲", 30],
    ]
    for i, row_data in enumerate(data4):
        row_num = 3 + i
        for j, val in enumerate(row_data):
            ws4.cell(row_num, j + 1, val)

    # 保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"测试 Excel 已创建: {output_path}")
    return output_path


if __name__ == "__main__":
    test_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_data", "test_config.xlsx")
    create_test_excel(test_path)
